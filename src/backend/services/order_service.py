from typing import List, Dict, Any
from sqlalchemy.orm import Session, joinedload
from fastapi import UploadFile
from datetime import datetime
import openpyxl
import tempfile
import os

from models.order import Order
from models.goods import Goods
from models.node import Node
from schemas.order import OrderCreate, OrderUpdate, OrderResponse, OrderImportResponse
from core.error_codes import (CODE_SUCCESS, CODE_PARAM_ERROR, CODE_INTERNAL_ERROR,
                             CODE_ORDER_NOT_FOUND, CODE_ORDER_STATUS_NOT_ALLOWED,
                             CODE_NODE_NOT_FOUND)


class OrderService:
    """订单服务"""

    @staticmethod
    async def create_order(order_create: OrderCreate, db: Session) -> Dict[str, Any]:
        """创建订单"""
        try:
            # 1. 校验destination_node_code是否存在
            dest_node = db.query(Node).filter(Node.node_code == order_create.destination_node_code).first()
            if not dest_node:
                return {"code": CODE_NODE_NOT_FOUND, "message": "目的地节点不存在", "data": None}

            # 2. 生成order_code
            import time
            order_code = f"O{int(time.time() * 1000)}"

            # 3. 创建Order记录
            order = Order(
                order_code=order_code,
                destination_node_id=dest_node.id,
                time_window=order_create.time_window,
                status="pending"
            )
            db.add(order)
            db.flush()  # 获取order.id

            # 4. 创建Goods记录
            for idx, goods_item in enumerate(order_create.goods):
                goods_code = f"G{int(time.time() * 1000)}_{idx}"
                goods = Goods(
                    goods_code=goods_code,
                    order_id=order.id,
                    goods_name=goods_item.goods_name,
                    goods_type=goods_item.goods_type,
                    weight=goods_item.weight,
                    volume=goods_item.volume,
                    node_id=dest_node.id,  # 初始在目的地节点
                    status="pending_pack"
                )
                db.add(goods)

            db.commit()

            # 5. 返回响应
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code,
                    "destination_node_name": dest_node.name,
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods_count": len(order_create.goods),
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"创建订单失败: {str(e)}", "data": None}

    @staticmethod
    async def get_orders(page: int, page_size: int, status: str = None, db: Session = None) -> Dict[str, Any]:
        """获取订单列表"""
        try:
            # 1. 构建查询（使用joinedload预加载关联对象，减少N+1查询）
            query = db.query(Order).options(
                joinedload(Order.destination_node),
                joinedload(Order.goods)
            )
            if status:
                query = query.filter(Order.status == status)

            # 2. 分页
            total = query.count()
            orders = query.offset((page - 1) * page_size).limit(page_size).all()

            # 3. 构建响应
            items = []
            for order in orders:
                dest_node = order.destination_node
                goods_count = len(order.goods) if order.goods else 0
                items.append({
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods_count": goods_count,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                })

            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "items": items,
                    "total": total,
                    "page": page,
                    "page_size": page_size
                }
            }
        except Exception as e:
            return {"code": CODE_INTERNAL_ERROR, "message": f"获取订单列表失败: {str(e)}", "data": None}

    @staticmethod
    async def get_order(order_code: str, db: Session) -> Dict[str, Any]:
        """获取订单详情"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 获取目的地节点
            dest_node = db.query(Node).filter(Node.id == order.destination_node_id).first()

            # 3. 获取货物列表
            goods = db.query(Goods).filter(Goods.order_id == order.id).all()
            goods_list = []
            for g in goods:
                goods_list.append({
                    "goods_code": g.goods_code,
                    "goods_name": g.goods_name,
                    "goods_type": g.goods_type,
                    "weight": float(g.weight),
                    "volume": float(g.volume),
                    "status": g.status
                })

            # 4. 返回响应
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "goods": goods_list,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            return {"code": CODE_INTERNAL_ERROR, "message": f"获取订单详情失败: {str(e)}", "data": None}

    @staticmethod
    async def update_order(order_code: str, order_update: OrderUpdate, db: Session) -> Dict[str, Any]:
        """更新订单"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 校验订单状态（delivering/completed/exception不可修改）
            if order.status in ["delivering", "completed", "exception"]:
                return {"code": CODE_ORDER_STATUS_NOT_ALLOWED, "message": "订单状态不允许修改", "data": None}

            # 3. 更新字段
            if order_update.destination_node_code is not None:
                dest_node = db.query(Node).filter(Node.node_code == order_update.destination_node_code).first()
                if not dest_node:
                    return {"code": CODE_NODE_NOT_FOUND, "message": "目的地节点不存在", "data": None}
                order.destination_node_id = dest_node.id

            if order_update.time_window is not None:
                order.time_window = order_update.time_window

            order.updated_at = datetime.now()
            db.commit()

            # 4. 返回响应
            dest_node = db.query(Node).filter(Node.id == order.destination_node_id).first()
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "order_code": order.order_code,
                    "destination_node_code": dest_node.node_code if dest_node else "",
                    "destination_node_name": dest_node.name if dest_node else "",
                    "time_window": order.time_window,
                    "status": order.status,
                    "created_at": order.created_at.isoformat(),
                    "updated_at": order.updated_at.isoformat()
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"更新订单失败: {str(e)}", "data": None}

    @staticmethod
    async def delete_order(order_code: str, db: Session) -> Dict[str, Any]:
        """删除订单"""
        try:
            # 1. 查询Order
            order = db.query(Order).filter(Order.order_code == order_code).first()
            if not order:
                return {"code": CODE_ORDER_NOT_FOUND, "message": "订单不存在", "data": None}

            # 2. 校验订单状态（delivering/completed/exception不可删除）
            if order.status in ["delivering", "completed", "exception"]:
                return {"code": CODE_ORDER_STATUS_NOT_ALLOWED, "message": "订单状态不允许删除", "data": None}

            # 3. 删除订单（会级联删除货物）
            db.delete(order)
            db.commit()

            return {"code": CODE_SUCCESS, "message": "success", "data": None}
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"删除订单失败: {str(e)}", "data": None}

    @staticmethod
    async def import_orders(file: UploadFile, skip_errors: bool, db: Session) -> Dict[str, Any]:
        """批量导入订单"""
        try:
            # 1. 读取文件
            contents = await file.read()
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(contents)
                tmp_path = tmp.name

            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            os.unlink(tmp_path)  # 删除临时文件

            # 2. 解析表头
            headers = [cell.value for cell in ws[1]]

            # 3. 逐行校验并处理
            success_count = 0
            failed_count = 0
            failed_rows = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 解析行数据
                    row_data = dict(zip(headers, row))
                    destination_node_code = row_data.get("destination_node_code")
                    time_window = row_data.get("time_window")
                    goods_name = row_data.get("goods_name")
                    goods_type = row_data.get("goods_type")
                    weight = row_data.get("weight")
                    volume = row_data.get("volume")

                    # 校验必填字段
                    if not all([destination_node_code, time_window, goods_name, goods_type, weight, volume]):
                        raise ValueError("必填字段不能为空")

                    # 校验目的地节点是否存在
                    dest_node = db.query(Node).filter(Node.node_code == destination_node_code).first()
                    if not dest_node:
                        raise ValueError(f"目的地节点不存在: {destination_node_code}")

                    # 创建订单
                    import time
                    order_code = f"O{int(time.time() * 1000)}_{row_idx}"
                    order = Order(
                        order_code=order_code,
                        destination_node_id=dest_node.id,
                        time_window=time_window,
                        status="pending"
                    )
                    db.add(order)
                    db.flush()

                    # 创建货物
                    goods_code = f"G{int(time.time() * 1000)}_{row_idx}"
                    goods = Goods(
                        goods_code=goods_code,
                        order_id=order.id,
                        goods_name=goods_name,
                        goods_type=goods_type,
                        weight=weight,
                        volume=volume,
                        node_id=dest_node.id,
                        status="pending_pack"
                    )
                    db.add(goods)

                    success_count += 1
                except Exception as e:
                    failed_count += 1
                    failed_rows.append({
                        "row": row_idx,
                        "error": str(e)
                    })
                    if not skip_errors:
                        raise e

            db.commit()

            # 4. 返回结果
            return {
                "code": CODE_SUCCESS,
                "message": "success",
                "data": {
                    "success_count": success_count,
                    "failed_count": failed_count,
                    "failed_rows": failed_rows
                }
            }
        except Exception as e:
            db.rollback()
            return {"code": CODE_INTERNAL_ERROR, "message": f"导入订单失败: {str(e)}", "data": None}

